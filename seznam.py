import random
import time

import openpyxl
from bs4 import BeautifulSoup
from selenium import webdriver

from googleparser10 import get_driver, check_next, proxy, get_one_page


def get_data(key, month, one_day, year):

    k = 10

    proxy = random.choice(proxies)
    PROXY = f'http://indy71070@Tomt4485JG:{proxy}'

    url = f'https://www.google.com/search?q={key}&newwindow=1&tbs=cdr:1,cd_min:{month}/{one_day}/{year + 2018},cd_max:{month}/{one_day}/{year + 2018},&tbm=nws&sxsrf=APq-WBt2I4CFEqEPhkSL_wl_cb3g_1a4Tw:1646918436061&ei=JPspYuKrA4L7rgS0rb2gBQ&start=0&sa=N&ved=2ahUKEwjizeui0bv2AhWCvYsKHbRWD1QQ8tMDegQIARA9&biw=2195&bih=1124&dpr=1.75'
    driver = get_driver(PROXY)
    driver.get(url)
    headers, links, texts = get_one_page(driver)
    while check_next(driver):
        url = f'https://www.google.com/search?q={key}&newwindow=1&tbs=cdr:1,cd_min:{month}/{one_day}/{year + 2018},cd_max:{month}/{one_day}/{year + 2018},&tbm=nws&sxsrf=APq-WBt2I4CFEqEPhkSL_wl_cb3g_1a4Tw:1646918436061&ei=JPspYuKrA4L7rgS0rb2gBQ&start={k}&sa=N&ved=2ahUKEwjizeui0bv2AhWCvYsKHbRWD1QQ8tMDegQIARA9&biw=2195&bih=1124&dpr=1.75'
        driver.get(url)
        header, link, text = get_one_page(driver)
        headers += header
        links += link
        texts += text
        k += 10

    print(len(headers))
    print(headers, links, texts)
    return headers, links, texts

proxies = proxy()


thirty_one_days = [i for i in range(1, 32)]
thirty_days = [i for i in range(1, 31)]
february = [i for i in range(1, 29)]
m = {1: thirty_one_days, 2: february, 3: thirty_one_days, 4: thirty_days,
     5: thirty_one_days, 6: thirty_days, 7: thirty_one_days, 8: thirty_one_days,
     9: thirty_days, 10: thirty_one_days, 11: thirty_days, 12: thirty_one_days}

pathfile = r'C:\Users\khali\OneDrive\Рабочий стол\protests.xlsx' #Указать путь к файлу
wb = openpyxl.load_workbook(pathfile)
sheet = wb.active

keys = ['protests']
k = 1


for key in keys:
    for year in range(4):
        for months, day in m.items():
            for one_day in day:
                data = get_data(key, months, one_day, year)
                headers = data[0]
                links = data[1]
                texts = data[2]

                for i in range(len(headers)):
                    sheet.cell(column=1, row=k, value=k)
                    sheet.cell(column=2, row=k, value=key)
                    sheet.cell(column=3, row=k, value=i + 1)
                    sheet.cell(column=4, row=k, value=headers[i])
                    sheet.cell(column=5, row=k, value=texts[i])
                    sheet.cell(column=6, row=k, value=links[i])
                    sheet.cell(column=7, row=k, value=str(one_day) + '.' + str(months) + '.' + str(year + 2018))
                    k += 1
            wb.save(pathfile)